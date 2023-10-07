
from openpyxl import *
from tkinter import *

#mở file excel ở thư mục...
wb = load_workbook('D:\\Nam_3\\1\python\\2113005_lab4\\demopython.xlsx')
#tạo đối tượng cho trang excel
sheet = wb.active
#thiết kế lại cột trong excel
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

# ghi dữ liệu vào excel tại vị trí(giao dữa cột và dòng)
    sheet.cell(row = 1, column = 1).value = "tên"
    sheet.cell(row = 1, column = 2).value = "niên khóa"
    sheet.cell(row = 1, column = 3).value = "học kì"
    sheet.cell(row = 1, column = 4).value = "mssv"
    sheet.cell(row = 1, column = 5).value = "số điện thoại"
    sheet.cell(row = 1, column = 6).value = "email"
    sheet.cell(row = 1, column = 7).value = "địa chỉ"

#đặt tiêu điểm con trỏ
def focus1(event):
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

    
def insert():
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):		
        print("empty input")
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column


        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
#lưu file
        wb.save('D:\\Nam_3\\1\python\\2113005_lab4\\demopython.xlsx')
#đặt tiêu điểm vào ô name_field
        name_field.focus_set()
#gọi hàm clear
        clear()

#chạy code
if __name__ == "__main__":
#tạo cửa sổ GUI
    root =Tk()
# đặt màu cho nền của cửa sổ GUI
    root.configure(background = "light green")
#cài đặt tiêu đề cho cửa sổ GUI
    root.title("resistration form")
#thiết lập cấu hình cho GUI
    root.geometry("500x300")
    excel
#tạo một cái nhãn trong form
    heading = Label(root, text = "Form", bg = "light green")
#tạo tên nhãn
    name = Label(root, text="Name", bg="light green")
#tạo cấu hình cho label
    course = Label(root, text="Course", bg="light green")
#tạo một nhãn học kì
    sem = Label(root, text="Semester", bg="light green")
#tạo nhãn mẫu số
    form_no = Label(root, text="Form No.", bg="light green")
#tạo kết nối với nhãn
    contact_no = Label(root, text="Contact No.", bg="light green")
#tạo một email có id nhãn
    email_id = Label(root, text="Email id", bg="light green")
#tạo một địa chỉ cho nhãn
    address = Label(root, text="Address", bg="light green")

    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)


    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)

    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    excel()

    submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
    submit.grid(row=8, column=1)

    root.mainloop()